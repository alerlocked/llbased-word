"""
Excel导出服务 - 支持每页一个Sheet的导出格式

将PDF表格导出为Excel格式，每页对应一个Sheet，
便于工艺文件的对照查看和编辑。

特性：
- 按页码分组导出到不同Sheet
- 支持合并同页多个表格
- 可选添加元数据信息
- 支持单Sheet和分Sheet两种模式
"""
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
from datetime import datetime

from app.shared.logging import get_logger

logger = get_logger(__name__)


class ExcelExportService:
    """
    Excel导出服务

    将PDF表格导出为Excel格式，每页对应一个Sheet
    """

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化Excel导出服务

        Args:
            config: 配置参数
                - include_metadata: 是否包含元数据
                - sheet_name_prefix: Sheet名称前缀
        """
        self.config = config or {}
        self.include_metadata = self.config.get("include_metadata", True)
        self.sheet_name_prefix = self.config.get("sheet_name_prefix", "第")

        logger.info("excel_export_service_initialized",
                   include_metadata=self.include_metadata)

    def export_tables_to_excel(
        self,
        tables: List[Any],
        output_path: Union[str, Path],
        group_by_page: bool = True
    ) -> Dict[str, Any]:
        """
        导出表格到Excel，按页分Sheet

        Args:
            tables: 表格列表（ExtractedTable或字典）
            output_path: 输出路径
            group_by_page: 是否按页分组到不同Sheet

        Returns:
            导出结果信息
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError as e:
            logger.error("excel_dependencies_missing", error=str(e))
            raise ImportError(
                "Excel导出需要安装openpyxl: pip install openpyxl"
            )

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if not tables:
            logger.warning("no_tables_to_export")
            return {
                "filepath": str(output_path),
                "sheets_created": 0,
                "total_tables": 0,
                "error": "没有表格可导出"
            }

        if group_by_page:
            return self._export_by_page(tables, output_path)
        else:
            return self._export_single_sheet(tables, output_path)

    def _export_by_page(
        self,
        tables: List[Any],
        output_path: Path
    ) -> Dict[str, Any]:
        """
        按页导出到不同Sheet

        Args:
            tables: 表格列表
            output_path: 输出路径

        Returns:
            导出结果信息
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            raise

        # 按页码分组
        page_tables: Dict[int, List[Any]] = {}
        for table in tables:
            page_num = self._get_page_number(table)
            if page_num not in page_tables:
                page_tables[page_num] = []
            page_tables[page_num].append(table)

        # 创建工作簿
        wb = Workbook()
        # 移除默认Sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        sheets_created = 0

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        for page_num in sorted(page_tables.keys()):
            page_table_list = page_tables[page_num]

            # 创建Sheet
            sheet_name = f"{self.sheet_name_prefix}{page_num + 1}页"
            ws = wb.create_sheet(title=sheet_name)

            current_row = 1

            for table_idx, table in enumerate(page_table_list):
                # 获取表格数据
                rows = self._get_table_rows(table)
                headers = self._get_table_headers(table)

                if not rows:
                    continue

                # 如果不是第一个表格，添加分隔
                if table_idx > 0:
                    current_row += 2  # 添加空行分隔

                # 写入表头
                if headers:
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = center_alignment
                    current_row += 1

                # 写入数据行
                for row in rows:
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = thin_border
                    current_row += 1

                # 添加表格元数据
                if self.include_metadata:
                    current_row += 1
                    meta_row = current_row
                    metadata = self._get_table_metadata(table)
                    if metadata:
                        ws.cell(row=meta_row, column=1, value=f"[置信度: {metadata.get('confidence', 'N/A')}]")
                        ws.cell(row=meta_row, column=2, value=f"[方法: {metadata.get('method', 'N/A')}]")
                        current_row += 1

            sheets_created += 1

        # 如果没有创建任何Sheet，创建一个空的
        if sheets_created == 0:
            ws = wb.create_sheet(title="无数据")
            ws.cell(row=1, column=1, value="没有可导出的表格数据")
            sheets_created = 1

        # 保存文件
        wb.save(output_path)

        result = {
            "filepath": str(output_path),
            "sheets_created": sheets_created,
            "total_tables": len(tables),
            "pages_included": list(page_tables.keys()),
            "exported_at": datetime.now().isoformat()
        }

        logger.info("excel_export_completed",
                   filepath=str(output_path),
                   sheets=sheets_created,
                   tables=len(tables))

        return result

    def _export_single_sheet(
        self,
        tables: List[Any],
        output_path: Path
    ) -> Dict[str, Any]:
        """
        导出到单个Sheet

        Args:
            tables: 表格列表
            output_path: 输出路径

        Returns:
            导出结果信息
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
        except ImportError:
            raise

        wb = Workbook()
        ws = wb.active
        ws.title = "全部表格"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        for table in tables:
            page_num = self._get_page_number(table)
            rows = self._get_table_rows(table)
            headers = self._get_table_headers(table)

            # 添加页面标记
            ws.cell(row=current_row, column=1, value=f"=== {self.sheet_name_prefix}{page_num + 1}页 ===")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="0066CC")
            current_row += 1

            # 写入表头
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                current_row += 1

            # 写入数据行
            for row in rows:
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                current_row += 1

            # 添加空行分隔
            current_row += 2

        wb.save(output_path)

        return {
            "filepath": str(output_path),
            "sheets_created": 1,
            "total_tables": len(tables)
        }

    def _get_page_number(self, table: Any) -> int:
        """获取表格的页码"""
        if hasattr(table, 'page_number'):
            return table.page_number
        elif isinstance(table, dict):
            return table.get('page_number', 0)
        return 0

    def _get_table_rows(self, table: Any) -> List[List[str]]:
        """获取表格的行数据"""
        if hasattr(table, 'rows'):
            return table.rows
        elif hasattr(table, 'data_rows') and table.data_rows:
            return table.data_rows
        elif isinstance(table, dict):
            return table.get('rows', []) or table.get('data_rows', [])
        return []

    def _get_table_headers(self, table: Any) -> Optional[List[str]]:
        """获取表格的表头"""
        if hasattr(table, 'headers'):
            return table.headers
        elif isinstance(table, dict):
            return table.get('headers')
        return None

    def _get_table_metadata(self, table: Any) -> Dict[str, Any]:
        """获取表格的元数据"""
        metadata = {}

        if hasattr(table, 'confidence_score'):
            metadata['confidence'] = f"{table.confidence_score:.2f}"
        elif isinstance(table, dict):
            conf = table.get('confidence_score') or table.get('confidence')
            if conf:
                metadata['confidence'] = f"{conf:.2f}" if isinstance(conf, float) else str(conf)

        if hasattr(table, 'extraction_method'):
            metadata['method'] = table.extraction_method
        elif isinstance(table, dict):
            metadata['method'] = table.get('extraction_method', 'unknown')

        if hasattr(table, 'parser_used'):
            metadata['parser'] = table.parser_used.value if hasattr(table.parser_used, 'value') else str(table.parser_used)
        elif isinstance(table, dict):
            parser = table.get('parser_used')
            if parser:
                metadata['parser'] = parser

        return metadata

    def export_to_excel_with_template(
        self,
        tables: List[Any],
        output_path: Union[str, Path],
        template_type: str = "process_document"
    ) -> Dict[str, Any]:
        """
        使用模板导出Excel

        Args:
            tables: 表格列表
            output_path: 输出路径
            template_type: 模板类型 (process_document, material_list, etc.)

        Returns:
            导出结果信息
        """
        # 根据模板类型选择导出方式
        if template_type == "process_document":
            return self._export_by_page(tables, Path(output_path))
        elif template_type == "material_list":
            # 材料表单Sheet导出
            return self._export_single_sheet(tables, Path(output_path))
        else:
            return self.export_tables_to_excel(tables, output_path, group_by_page=True)
